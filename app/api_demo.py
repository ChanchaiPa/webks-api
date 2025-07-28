from asyncio import run
from datetime import datetime
import io
from fastapi import APIRouter, File, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from app.cachedata import Cache

router = APIRouter()


@router.get("/demo/mock_data")
def mock_data(wording: str = ""):
    mock_data = run( Cache.getMockData() )
    if wording== "":
        return mock_data

    _wording = wording.lower()
    new_data = []
    for data in mock_data:
        _fname = data["first_name"].lower()
        if len(wording)>1 and _fname.startswith(_wording) :
            if len(new_data)>10:
                break
            else:
                new_data.append(data)  
    if len(new_data)>10:
        return new_data            

    for data in mock_data:
        _fname = data["first_name"].lower()
        if len(wording)>1 and _fname.startswith( _wording )==False and _wording in _fname :
            if len(new_data)>10:
                break
            else:
                new_data.append(data)
    return new_data


@router.post("/demo/excel")
async def mock_data(file: UploadFile|None = File(...)):
    file_content = await file.read()
    excel_file_object = io.BytesIO(file_content)
    workbook: Workbook = load_workbook(excel_file_object)
    sheet0 = workbook.worksheets[0]

    for i, row in enumerate(sheet0.iter_rows(), start=1):
        if  i==1: #header
            continue
        for j, cell in enumerate(row, start=1):
            isError = False
            val = str(cell.value).strip()
            if j==4 or j==5:
                try :
                    dateObj: datetime = datetime.strptime(val, "%d/%m/%Y")
                    if dateObj.year > 2099:
                        isError = True
                except ValueError as e:
                    isError = True
                    print(str(e) + "   '"+val+"'")
                if isError:
                    cell.font = Font(name='Arial', size=12, bold=True, italic=True, color="FF0000") # Red color
                    print(str(i) + "," + str(j) + ") " + val + " ERROR")
                    sheet0.cell(row=i, column=15, value="Invalid Value").font = Font(name='Arial', size=12, bold=True,
                                                                                    italic=True, color="FF0000")
    response_stream = io.BytesIO()
    workbook.save(response_stream)
    response_stream.seek(0)
    return StreamingResponse(
        response_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=ERROR-"+file.filename
        }
    )